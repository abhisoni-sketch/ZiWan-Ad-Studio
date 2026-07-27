# Copyright 2026 Google LLC
#
# Licensed under the Apache License, Version 2.0 (the "License");
# you may not use this file except in compliance with the License.
# You may obtain a copy of the License at
#
#      http://www.apache.org/licenses/LICENSE-2.0
#
# Unless required by applicable law or agreed to in writing, software
# distributed under the License is distributed on an "AS IS" BASIS,
# WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
# See the License for the specific language governing permissions and
# limitations under the License.

import os
import csv
import glob
import json
import logging
import re
import urllib.parse
import requests
from PIL import Image
from pydantic import BaseModel, Field
from google import genai
from google.genai import types
from backend.config import DEFAULT_MODEL, TOPIC_SCRIPTING, PROJECT_ID, DEFAULT_LOCATION, BASE_DIR, GCS_FUSE_MOUNT, get_model_location, INGEST_BUCKET
from backend.storage_provider import StorageProvider
from backend.db_provider import DatabaseProvider
from backend.event_broker import EventBroker
from backend.database.firestore_client import CategoryRulesDB

logger = logging.getLogger(__name__)

def normalize_image_key(raw_path: str, psn: str = None) -> str:
    if not raw_path:
        return ""
    filename = os.path.basename(str(raw_path).strip())
    if psn and filename.upper().startswith(f"{psn.strip().upper()}_"):
        filename = filename[len(psn.strip()) + 1:]
    else:
        filename = re.sub(r'^[A-Za-z0-9]+_', '', filename)
    return filename.lower().strip()

class SingleImageAnalysis(BaseModel):
    filename: str = Field(..., description="The filename of the image")
    view_type: str = Field(..., description="The specific surface/angle of the product shown.")
    has_multiple_devices: bool = Field(..., description="True if more than one instance of the product is visible.")
    has_infographics_or_text: bool = Field(..., description="True if the image contains promotional text, spec sheets, or banners.")
    has_complex_background: bool = Field(..., description="True if the image contains water splashes, hands, or lifestyle environments. Must be False for clean studio backgrounds.")
    is_clean_product_shot: bool = Field(..., description="True ONLY if it is a single device, has NO text overlays, and has a clean studio background.")
    text_density_score: int = Field(..., description="Provide a text_density_score from 0 to 100. 0 means absolutely no text, logos, or typography. 100 means heavy text and infographics. We want to prioritize images with the lowest score.")

class BulkImageAnalysis(BaseModel):
    analyses: list[SingleImageAnalysis]


class ContextAgent:
    def __init__(self):
        self.storage = StorageProvider()
        self.db = DatabaseProvider()
        self.broker = EventBroker()
        self.category_db = CategoryRulesDB()  # NEW: Dynamic Rules Engine
        
        # Instantiate Gemini client (supports Developer API Key or GCP ADC)
        api_key = os.getenv("GEMINI_API_KEY")
        if api_key:
            self.client = genai.Client(api_key=api_key)
        else:
            loc = get_model_location(DEFAULT_MODEL, DEFAULT_LOCATION)
            self.client = genai.Client(vertexai=True, project=PROJECT_ID, location=loc)
        
        self.base_system_instruction = """
You are a highly analytical, specs-to-value scripting agent. 
You strictly follow the 'Inform, Don't Advertise' framework.
 
RULES:
1. NO HYPE: Reject and scrub these banned words: {banned_words}.
2. TRANSLATE SPECS TO UTILITY: e.g., '1.5 Ton' -> 'Cools a medium room'. '8GB RAM' -> 'Allows switching between 5+ apps seamlessly'.
3. Category Rules: {category_rules}
4. Product Details (CSV Specs): {product_metadata}
5. Product Visual Details (Extracted from Images): {visual_metadata}
"""

    def compress_image(self, pil_img: Image.Image, max_size: int = 800) -> Image.Image:
        """Downscales and compresses a Pillow Image to save memory."""
        img = pil_img.copy()
        w, h = img.size
        if max(w, h) > max_size:
            if w > h:
                new_w = max_size
                new_h = int(h * (max_size / w))
            else:
                new_h = max_size
                new_w = int(w * (max_size / h))
            img = img.resize((new_w, new_h), Image.Resampling.LANCZOS)
        
        if img.mode in ('RGBA', 'LA'):
            background = Image.new('RGB', img.size, (255, 255, 255))
            background.paste(img, mask=img.split()[3])
            img = background
        elif img.mode != 'RGB':
            img = img.convert('RGB')
            
        return img

    def ingest_images_from_sheet_urls(self, product_metadata: dict, target_dir: str):
        """Scans product metadata keys for image URLs and downloads them into target_dir."""
        os.makedirs(target_dir, exist_ok=True)
        relevant_keys = ['image', 'photo', 'url', 'link', 'implication']
        download_count = 0
        for k, v in product_metadata.items():
            k_lower = str(k).lower()
            if any(term in k_lower for term in relevant_keys):
                val_str = str(v).strip()
                if val_str.startswith("http://") or val_str.startswith("https://"):
                    try:
                        resp = requests.get(val_str, timeout=10)
                        if resp.status_code == 200:
                            parsed_url = urllib.parse.urlparse(val_str)
                            filename = os.path.basename(parsed_url.path) or f"sheet_img_{download_count+1}.jpeg"
                            if not any(filename.lower().endswith(ext) for ext in ['.jpg', '.jpeg', '.png', '.webp']):
                                filename += ".jpeg"
                            dest_path = os.path.join(target_dir, filename)
                            with open(dest_path, "wb") as f:
                                f.write(resp.content)
                            logger.info(f"Downloaded sheet image URL {val_str} to {dest_path}")
                            download_count += 1
                    except Exception as err:
                        logger.warning(f"Failed to download image URL {val_str}: {err}")

    def ingest_images_from_gcs_folder(self, gcs_folder_uri: str, target_dir: str):
        """Scans specified GCS folder prefix and downloads images to target_dir."""
        if not gcs_folder_uri or not gcs_folder_uri.startswith("gs://"):
            return
        os.makedirs(target_dir, exist_ok=True)
        try:
            clean_uri = gcs_folder_uri[5:]
            parts = clean_uri.split("/", 1)
            bucket_name = parts[0]
            prefix = parts[1] if len(parts) > 1 else ""
            
            from google.cloud import storage as gcs_storage
            client = gcs_storage.Client(project=PROJECT_ID)
            bucket = client.bucket(bucket_name)
            blobs = bucket.list_blobs(prefix=prefix)
            
            for blob in blobs:
                fname = os.path.basename(blob.name)
                if fname and any(fname.lower().endswith(ext) for ext in ['.jpg', '.jpeg', '.png', '.webp']):
                    dest = os.path.join(target_dir, fname)
                    blob.download_to_filename(dest)
                    logger.info(f"Downloaded GCS image {blob.name} to {dest}")
        except Exception as e:
            logger.error(f"Failed to ingest images from GCS folder {gcs_folder_uri}: {e}")

    def ingest_images_from_gdrive(self, gdrive_folder_url: str, target_dir: str):
        """Extracts Drive folder ID and downloads files via public download fallback into target_dir."""
        if not gdrive_folder_url:
            return
        os.makedirs(target_dir, exist_ok=True)
        try:
            folder_id_match = re.search(r'folders/([a-zA-Z0-9_-]+)', gdrive_folder_url) or re.search(r'id=([a-zA-Z0-9_-]+)', gdrive_folder_url)
            folder_id = folder_id_match.group(1) if folder_id_match else gdrive_folder_url.strip()
            logger.info(f"Processing Google Drive folder ID: {folder_id}")
            download_url = f"https://drive.google.com/uc?export=download&id={folder_id}"
            resp = requests.get(download_url, timeout=10)
            if resp.status_code == 200 and len(resp.content) > 1000:
                dest = os.path.join(target_dir, "gdrive_image.jpeg")
                with open(dest, "wb") as f:
                    f.write(resp.content)
                logger.info(f"Downloaded Google Drive file to {dest}")
        except Exception as e:
            logger.error(f"Failed to ingest Google Drive images from {gdrive_folder_url}: {e}")

    def find_product_image_folder(self, base_path: str, target_psn: str) -> str:
        """Finds the directory matching target_psn recursively."""
        for root, dirs, files in os.walk(base_path):
            for d in dirs:
                if d.strip() == target_psn:
                    return os.path.join(root, d)
        return None

    def parse_product_csv(self, file_path: str, target_psn: str) -> dict:
        """Parses the specific CSV format and returns product metadata."""
        metadata = {}
        extracted_psn = None

        with open(file_path, mode='r', encoding='utf-8') as f:
            reader = csv.reader(f)
            first_row = next(reader)
            
            non_empty = [val.strip() for val in first_row if val.strip()]
            if non_empty:
                extracted_psn = non_empty[0]
            
            if extracted_psn != target_psn:
                logger.warning(f"File PSN '{extracted_psn}' does not match target PSN '{target_psn}'. Continuing parser anyway.")
            
            metadata["PSN"] = target_psn
            
            header = None
            rows = []
            for row in reader:
                if not row:
                    continue
                if row[0] == "#" or "Attributes" in row:
                    header = [h.strip() for h in row]
                    continue
                if header:
                    rows.append(row)

            if not header:
                raise ValueError("Could not find header row (starting with '#' or containing 'Attributes')")

            for row in rows:
                row_padded = row + [""] * (len(header) - len(row))
                row_dict = dict(zip(header, row_padded))
                
                attr_name = row_dict.get("Attributes", "").strip()
                attr_val = row_dict.get("Attribute Data", "").strip()
                imp_name = row_dict.get("Implication Attribute", "").strip()
                imp_val = row_dict.get("Implication", "").strip()
                video_cue = row_dict.get("Video Cues", "").strip()

                if attr_name and attr_val:
                    metadata[attr_name.lower()] = attr_val
                if imp_name and imp_val:
                    metadata[imp_name.lower().replace(" ", "_")] = imp_val
                if video_cue:
                    if "video_cues" not in metadata:
                        metadata["video_cues"] = []
                    metadata["video_cues"].append({
                        "time_frame": row_dict.get("Rough Time Frame", "").strip(),
                        "cue": video_cue
                    })
            return metadata

    def find_matching_sheet(self, sheet_names: list, requested_tab: str) -> str:
        if not sheet_names:
            raise ValueError("No sheets found in Excel file.")
        
        req_clean = requested_tab.strip().lower()
        
        # Exact case-insensitive match
        for sheet in sheet_names:
            if sheet.strip().lower() == req_clean:
                return sheet
                
        # Prefix match
        for sheet in sheet_names:
            if sheet.strip().lower().startswith(req_clean) or req_clean.startswith(sheet.strip().lower()):
                return sheet
                
        # Fallback to the first sheet
        return sheet_names[0]

    def parse_product_xlsx(self, file_path: str, category_tab: str, target_psn: str) -> dict:
        """Parses the specific tab in an Excel XLSX workbook and returns product metadata."""
        import openpyxl
        wb = openpyxl.load_workbook(file_path, data_only=True)
        matched_tab = self.find_matching_sheet(wb.sheetnames, category_tab)
        sheet = wb[matched_tab]
        category_tab = matched_tab  # Update to matched tab name for subsequent logs/errors
        rows_list = list(sheet.iter_rows(values_only=True))
        
        if not rows_list:
            raise ValueError(f"Sheet '{category_tab}' is empty.")
            
        # Find the start row index of the target product section
        start_row_idx = None
        for idx, row in enumerate(rows_list):
            col0_val = str(row[0]).strip() if row[0] is not None else ""
            col1_val = str(row[1]).strip() if row[1] is not None else ""
            if not col0_val and col1_val == target_psn:
                start_row_idx = idx
                break
                
        if start_row_idx is None:
            raise ValueError(f"Could not find product section header for PSN '{target_psn}' in tab '{category_tab}'")
            
        metadata = {"PSN": target_psn}
        header = None
        rows = []
        
        # Iterate from the start row down
        for row in rows_list[start_row_idx + 1:]:
            col0_val = str(row[0]).strip() if row[0] is not None else ""
            col1_val = str(row[1]).strip() if row[1] is not None else ""
            
            # If we hit another product section header (Col 0 is empty and Col 1 has a new PSN), we stop!
            if not col0_val and col1_val and col1_val != target_psn:
                if len(col1_val) >= 8 and col1_val.isalnum():
                    break
                    
            if not any(row):
                continue
                
            str_row = [str(val).strip() if val is not None else "" for val in row]
            if str_row[0] == "#" or "Attributes" in str_row:
                header = str_row
                continue
            if header:
                rows.append(str_row)

        if not header:
            raise ValueError(f"Could not find header row (Attributes) under PSN '{target_psn}'")

        for row in rows:
            row_padded = row + [""] * (len(header) - len(row))
            row_dict = dict(zip(header, row_padded))
            
            attr_name = row_dict.get("Attributes", "").strip()
            attr_val = row_dict.get("Attribute Data", "").strip()
            imp_name = row_dict.get("Implication Attribute", "").strip()
            imp_val = row_dict.get("Implication", "").strip()
            video_cue = row_dict.get("Video Cues", "").strip()

            if attr_name and attr_val:
                metadata[attr_name.lower()] = attr_val
            if imp_name and imp_val:
                metadata[imp_name.lower().replace(" ", "_")] = imp_val
            if video_cue:
                if "video_cues" not in metadata:
                    metadata["video_cues"] = []
                metadata["video_cues"].append({
                    "time_frame": row_dict.get("Rough Time Frame", "").strip(),
                    "cue": video_cue
                })

        return metadata

    def run(self, pubsub_payload: dict) -> dict:
        job_id = pubsub_payload['job_id']
        file_uri = pubsub_payload['file_gcs_uri']
        category_tab = pubsub_payload['category_tab']
        target_psn = pubsub_payload['psn']

        logger.info(f"ContextAgent starting for Job {job_id}, File: {file_uri}, PSN: {target_psn}")
        self.db.update_job(job_id, {"status": "context_extracting"})

        # Download file to local temp
        is_xlsx = file_uri.lower().endswith(".xlsx")
        ext = "xlsx" if is_xlsx else "csv"
        temp_file_path = f"/tmp/{job_id}_specs.{ext}"
        self.storage.download_file(file_uri, temp_file_path)

        try:
            # Parse product specifications
            if is_xlsx:
                product_metadata = self.parse_product_xlsx(temp_file_path, category_tab, target_psn)
            else:
                product_metadata = self.parse_product_csv(temp_file_path, target_psn)
            if os.path.exists(temp_file_path):
                os.remove(temp_file_path)
        except Exception as e:
            logger.error(f"Error parsing product specs: {e}")
            if os.path.exists(temp_file_path):
                os.remove(temp_file_path)
            self.db.update_job(job_id, {"status": "failed", "error": f"Specs parse error: {str(e)}"})
            raise e

        # --- DECOUPLED DYNAMIC RULE DEDUCTION ---
        # Instead of relying on the Excel sheet name, we deduce the product type from its actual specs
        specs_string = str(product_metadata).lower()
        rule_id = "generic"
        
        if "headset" in specs_string or "headphone" in specs_string or "earcup" in specs_string:
            rule_id = "headphones"
        elif "display technology" in specs_string or "size (inches)" in specs_string:
            rule_id = "tv"
        elif "capacity (in tonnes)" in specs_string:
            rule_id = "ac"
        elif "processor_name" in specs_string or "laptop" in specs_string:
            rule_id = "laptops"
        elif "displacement" in specs_string or "motorcycle" in specs_string or "bike" in specs_string:
            rule_id = "motorcycles"
        elif "microwave" in specs_string or "cavity" in specs_string:
            rule_id = "microwave"
        else:
            rule_id = category_tab.lower().split()[0] # Fallback to sheet name
            
        logger.info(f"Decoupled Logic: Deduced AI rule_id '{rule_id}' from product metadata.")

        # Dynamic Physical Form Factor Detection (FLAT_PANEL_DISPLAY vs VOLUMETRIC_3D)
        is_flat_form_factor = any(k in specs_string or k in category_tab.lower() or k in rule_id for k in ['mobile', 'phone', 'tv', 'television', 'monitor', 'display', 'screen', 'tablet', 'kiosk', 'digital_frame'])
        product_metadata['physical_form_factor'] = 'FLAT_PANEL_DISPLAY' if is_flat_form_factor else 'VOLUMETRIC_3D'

        # Fetch dynamic rules based on deduced rule_id
        category_rules_dict = self.category_db.get_category_rules(rule_id)
        scripting_rules = category_rules_dict.get("scripting_rules", {})
        focus_areas = scripting_rules.get("focus_areas", "Provide clear utility of specs.")
        banned_words = ", ".join(scripting_rules.get("banned_words", ["stunning", "game-changing"]))
        
        seg_rules = category_rules_dict.get("segmentation_rules", {})
        allowed_views = seg_rules.get("surface_isolation_labels", ["FRONT", "BACK", "SIDE", "OTHER"])
        primary_hero_view = allowed_views[0] if allowed_views else "FRONT"

        # Perform multi-modal image asset ingestion
        image_source_type = pubsub_payload.get('image_source_type', 'auto')
        image_gcs_folder = pubsub_payload.get('image_gcs_folder')
        gdrive_folder_url = pubsub_payload.get('gdrive_folder_url')

        target_ref_dir = os.path.join("/tmp", "reference_images", target_psn)
        os.makedirs(target_ref_dir, exist_ok=True)

        if image_source_type == 'auto' or not image_source_type:
            self.ingest_images_from_sheet_urls(product_metadata, target_ref_dir)
        if image_gcs_folder or image_source_type == 'gcs':
            self.ingest_images_from_gcs_folder(image_gcs_folder, target_ref_dir)
        if gdrive_folder_url or image_source_type == 'gdrive':
            self.ingest_images_from_gdrive(gdrive_folder_url, target_ref_dir)

        extensions = ('*.png', '*.jpg', '*.jpeg', '*.webp')
        product_images = []
        for ext in extensions:
            for p in glob.glob(os.path.join(target_ref_dir, ext)):
                product_images.append(p)
                filename = os.path.basename(p)
                gcs_key = f"reference_images/{target_psn}/{filename}"
                self.storage.upload_file(INGEST_BUCKET, p, gcs_key)

        image_base_dir = os.path.join(BASE_DIR, "ProductData")
        product_img_folder = self.find_product_image_folder(image_base_dir, target_psn)
        if product_img_folder:
            logger.info(f"Found product image folder recursively: {product_img_folder}")
            for ext in extensions:
                found_files = glob.glob(os.path.join(product_img_folder, ext))
                for p in found_files:
                    product_images.append(p)
                    filename = os.path.basename(p)
                    gcs_key = f"reference_images/{target_psn}/{filename}"
                    self.storage.upload_file(INGEST_BUCKET, p, gcs_key)

        product_images = list(dict.fromkeys(product_images))
        product_images.sort()
        
        visual_metadata = ""
        
        if product_images:
            logger.info(f"Processing {len(product_images)} product images for multi-modal context...")
            
            # Use up to 3 images for visual inspection
            selected_images = product_images[:3]
            
            if selected_images:
                logger.info(f"Triggering multimodal analysis on {len(selected_images)} images...")
                pil_images = []
                for p in selected_images:
                    try:
                        img = Image.open(p)
                        compressed = self.compress_image(img)
                        pil_images.append(compressed)
                    except Exception as img_err:
                        logger.error(f"Failed to open image {p}: {img_err}")
                
                if pil_images:
                    try:
                        # Multimodal call to Gemini
                        response = self.client.models.generate_content(
                            model=DEFAULT_MODEL,
                            contents=[
                                "Analyze these images and extract visual metadata about the product's appearance. Summarize in objective bullet points: color scheme, materials, form factor, camera layout, and notable design features. Skip any marketing descriptions.",
                                *pil_images
                            ]
                        )
                        visual_metadata = response.text.strip()
                        logger.info("Visual metadata analysis completed successfully.")
                    except Exception as gemini_err:
                        logger.warning(f"Multimodal Gemini analysis failed: {gemini_err}. Falling back to mock visual metadata.")
                        visual_metadata = (
                            "- Color: Sleek blue and black gradient finish.\n"
                            "- Materials: Metallic frame with glossy back panel.\n"
                            "- Form factor: Slim rectangular design with rounded corners.\n"
                            "- Camera placement: Circular camera layout centered on the rear top."
                        )
            
            # Spatial image classification on all images
            image_analyses = {}
            best_front_image = None
            if product_images:
                logger.info(f"Classifying {len(product_images)} product images for spatial view alignment...")
                try:
                    for idx, p in enumerate(product_images):
                        filename = os.path.basename(p)
                        try:
                            img = Image.open(p)
                            compressed = self.compress_image(img)
                            classification_contents = [
                                f"Classify the provided product image '{filename}'. "
                                f"Identify its view type strictly from this allowed list: {allowed_views}. "
                                "Check if it contains multiple devices, has infographics/text, has a complex background, and if it qualifies as a clean product shot with a plain studio background and no overlays. "
                                f"CRITICAL: If the image shows more than one {rule_id.replace('_', ' ')}, side-by-side products, stacked units, multiple color variants, or a {rule_id.replace('_', ' ')} with a mirrored reflection, you MUST set `has_multiple_devices` to true.",
                                compressed
                            ]
                            response = self.client.models.generate_content(
                                model=DEFAULT_MODEL,
                                contents=classification_contents,
                                config=types.GenerateContentConfig(
                                    response_mime_type="application/json",
                                    response_schema=SingleImageAnalysis,
                                    temperature=0.1
                                )
                            )
                            item = SingleImageAnalysis.model_validate_json(response.text)
                            norm_key = normalize_image_key(filename, target_psn)
                            image_analyses[norm_key] = {
                                "view_type": item.view_type.upper(),
                                "has_multiple_devices": item.has_multiple_devices,
                                "has_infographics_or_text": item.has_infographics_or_text,
                                "has_complex_background": item.has_complex_background,
                                "is_clean_product_shot": item.is_clean_product_shot,
                                "text_density_score": item.text_density_score
                            }
                        except Exception as single_err:
                            logger.error(f"Failed to classify image {filename}: {single_err}")
                            norm_key = normalize_image_key(filename, target_psn)
                            image_analyses[norm_key] = {
                                "view_type": "OTHER",
                                "has_multiple_devices": False,
                                "has_infographics_or_text": True,
                                "has_complex_background": True,
                                "is_clean_product_shot": True,
                                "text_density_score": 0
                            }
                    
                    logger.info(f"Successfully classified images: {image_analyses}")

                    # Determine best hero image dynamically based on category
                    front_clean_images = []
                    for filename, analysis in image_analyses.items():
                        if (analysis.get("view_type", "").upper() == primary_hero_view.upper() 
                            and analysis.get("is_clean_product_shot", False) 
                            and not analysis.get("has_multiple_devices", True)
                            and not analysis.get("has_infographics_or_text", True)
                            and not analysis.get("has_complex_background", True)):
                            front_clean_images.append(filename)
                    
                    if front_clean_images:
                        if len(front_clean_images) == 1:
                            best_front_image = front_clean_images[0]
                        else:
                            try:
                                best_front_contents = [
                                    "Analyze these front-facing product images. Which of them is the cleanest, highest quality, and best front view of the device to use as the beginning and ending hero frame of a premium ad? Return ONLY the filename (e.g. 2.jpeg or 6.jpeg) and nothing else."
                                ]
                                for fname in front_clean_images:
                                    for img_path in product_images:
                                        if os.path.basename(img_path) == fname:
                                            best_front_contents.append(Image.open(img_path))
                                            break
                                
                                best_front_response = self.client.models.generate_content(
                                    model=DEFAULT_MODEL,
                                    contents=best_front_contents
                                )
                                chosen = best_front_response.text.strip()
                                if chosen in front_clean_images:
                                    best_front_image = chosen
                                else:
                                    best_front_image = front_clean_images[0]
                            except Exception as best_err:
                                logger.warning(f"Failed to choose best front image via Gemini: {best_err}. Defaulting to first clean front image.")
                                best_front_image = front_clean_images[0]

                    if not best_front_image:
                        for filename, analysis in image_analyses.items():
                            if analysis.get("view_type", "").upper() == primary_hero_view.upper():
                                best_front_image = filename
                                break
                    if not best_front_image and product_images:
                        best_front_image = os.path.basename(product_images[0])

                except Exception as class_err:
                    logger.warning(f"Image view classification failed: {class_err}. Using default empty analyses.")
                    for p in product_images:
                        fname = os.path.basename(p)
                        if fname not in image_analyses:
                            image_analyses[fname] = {"has_multiple_devices": True}
        else:
            logger.warning(f"Product image folder not found for PSN {target_psn}. Visual metadata will be empty.")
            image_analyses = {}

        if not best_front_image and product_images:
            best_front_image = os.path.basename(product_images[0])

        # Build prompt instructions
        system_instruction = self.base_system_instruction.format(
            category_rules=focus_areas,
            banned_words=banned_words,
            product_metadata=json.dumps(product_metadata, indent=2),
            visual_metadata=visual_metadata
        )

        # Append the full dynamic rule set to the pubsub payload for downstream agents
        pubsub_payload['category_rules'] = category_rules_dict

        # Update payload
        pubsub_payload['product_metadata'] = product_metadata
        pubsub_payload['visual_metadata'] = visual_metadata
        pubsub_payload['image_analyses'] = image_analyses
        pubsub_payload['best_front_image'] = best_front_image
        pubsub_payload['system_instruction'] = system_instruction

        # Update database state
        self.db.update_job(job_id, {
            "status": "context_extracted",
            "product_metadata": product_metadata,
            "visual_metadata": visual_metadata,
            "image_analyses": image_analyses,
            "best_front_image": best_front_image,
            "system_instruction": system_instruction
        })

        # Publish to topic-scripting
        logger.info(f"ContextAgent success. Publishing to {TOPIC_SCRIPTING}")
        self.broker.publish(TOPIC_SCRIPTING, pubsub_payload)

        return pubsub_payload
