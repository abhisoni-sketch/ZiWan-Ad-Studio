# Copyright 2026 Google LLC
#
# Licensed under the Apache License, Version 2.0 (the "License");
# you may not use this file except in compliance with the License.
# You may obtain a copy of the License at
#
#      http://www.apache.org/licenses/LICENSE-2.0
#
# Unless required by applicable law or agreed to in writing, software
# distributed under the License is distributed on an "AS IS" BASIS,
# WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
# See the License for the specific language governing permissions and
# limitations under the License.

import os
import uuid
import logging
import openpyxl
from backend.storage_provider import StorageProvider
from backend.db_provider import DatabaseProvider
from backend.event_broker import EventBroker
from backend.config import TOPIC_CONTEXT_AGENT

logger = logging.getLogger(__name__)

class BatchIngestionAgent:
    def __init__(self):
        self.storage = StorageProvider()
        self.db = DatabaseProvider()
        self.broker = EventBroker()

    def run(self, pubsub_payload: dict) -> dict:
        batch_id = pubsub_payload.get('batch_id', f"batch-{uuid.uuid4().hex[:8]}")
        file_uri = pubsub_payload['file_gcs_uri']
        model_target = pubsub_payload.get('model_target', 'gemini-omni-flash-preview')
        voice_name = pubsub_payload.get('voice_name')
        language_code = pubsub_payload.get('language_code')
        
        logger.info(f"BatchIngestionAgent starting for Batch {batch_id}, File: {file_uri}")
        
        if not file_uri.lower().endswith(".xlsx"):
            logger.error("Batch processing currently only supports .xlsx files.")
            return {"error": "Only .xlsx supported for batch."}

        temp_file_path = f"/tmp/{batch_id}_batch.xlsx"
        self.storage.download_file(file_uri, temp_file_path)

        try:
            wb = openpyxl.load_workbook(temp_file_path, data_only=True)
            total_jobs = 0
            
            # Iterate through all sheets (each sheet represents a category tab)
            for sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]
                rows_list = list(sheet.iter_rows(values_only=True))
                
                for row in rows_list:
                    col0_val = str(row[0]).strip() if row[0] is not None else ""
                    col1_val = str(row[1]).strip() if row[1] is not None else ""
                    
                    # Detect PSN Row: Col 0 is empty, Col 1 has a long alphanumeric string (the Serial Number)
                    if not col0_val and len(col1_val) >= 8 and col1_val.isalnum():
                        target_psn = col1_val
                        job_id = f"job-{uuid.uuid4().hex[:8]}"
                        logger.info(f"Batch {batch_id}: Found PSN {target_psn} in category '{sheet_name}'. Dispatching Job {job_id}.")
                        
                        payload = {
                            "job_id": job_id,
                            "batch_id": batch_id,
                            "file_gcs_uri": file_uri,
                            "psn": target_psn,
                            "category_tab": sheet_name,
                            "model_target": model_target,
                            "voice_name": voice_name,
                            "language_code": language_code,
                            "video_specs": {"resolution": "1080p", "aspect_ratio": "16:9"}
                        }
                        
                        # Write initial job status to database
                        self.db.create_job(job_id, {
                            "status": "ingesting",
                            "batch_id": batch_id,
                            "psn": target_psn,
                            "category": sheet_name,
                            "model_target": model_target,
                            "file_gcs_uri": file_uri
                        })
                        
                        # Publish to Context Agent topic to fan-out the job
                        self.broker.publish(TOPIC_CONTEXT_AGENT, payload)
                        total_jobs += 1

        except Exception as e:
            logger.error(f"Error parsing batch spreadsheet: {e}")
            raise e
        finally:
            if os.path.exists(temp_file_path):
                os.remove(temp_file_path)

        logger.info(f"Batch {batch_id} complete. {total_jobs} jobs dispatched.")
        return {"batch_id": batch_id, "jobs_dispatched": total_jobs}
